from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload
from typing import Optional, List
from datetime import datetime, date
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.models.assessment import (
    AssessmentTemplate, AssessmentDimension, AssessmentQuestion,
    CustomerAssessment, AssessmentResponse, AssessmentStatus,
    AssessmentResponseAudit, CustomerAssessmentTarget, AssessmentRecommendation
)
from app.models.mapping import RoadmapRecommendation
from app.schemas.assessment import (
    AssessmentTemplateCreate, AssessmentTemplateUpdate, AssessmentTemplateResponse,
    AssessmentTemplateDetailResponse, AssessmentTemplateListResponse,
    CustomerAssessmentCreate, CustomerAssessmentUpdate, CustomerAssessmentResponse,
    CustomerAssessmentDetailResponse, CustomerAssessmentListResponse,
    AssessmentAnswerCreate, AssessmentAnswerResponse, AssessmentAnswerWithQuestion,
    BatchResponseSubmit, AssessmentHistoryResponse, AssessmentComparison,
    ExcelUploadResult, ExcelResponseUploadResult,
    AssessmentAnswerUpdate, AssessmentAuditEntry, AssessmentAuditListResponse,
    TargetCreate, TargetUpdate, TargetResponse, TargetListResponse,
    DimensionGap, GapAnalysisResponse,
    FlowNode, FlowLink, FlowVisualizationResponse,
    AssessmentRecommendationCreate, AssessmentRecommendationUpdate,
    AssessmentRecommendationResponse, AssessmentRecommendationListResponse
)

router = APIRouter()


# ============================================================
# TEMPLATE ENDPOINTS
# ============================================================

@router.get("/templates", response_model=AssessmentTemplateListResponse)
async def list_templates(
    db: AsyncSession = Depends(get_db),
    active_only: bool = Query(False, description="Only show active templates"),
):
    """List all assessment templates."""
    query = select(AssessmentTemplate)

    if active_only:
        query = query.where(AssessmentTemplate.is_active == True)

    query = query.order_by(AssessmentTemplate.created_at.desc())
    query = query.options(selectinload(AssessmentTemplate.created_by))

    result = await db.execute(query)
    templates = result.scalars().all()

    return AssessmentTemplateListResponse(
        items=[AssessmentTemplateResponse.model_validate(t) for t in templates],
        total=len(templates)
    )


@router.get("/templates/active", response_model=Optional[AssessmentTemplateDetailResponse])
async def get_active_template(db: AsyncSession = Depends(get_db)):
    """Get the currently active assessment template with all questions."""
    query = select(AssessmentTemplate).where(
        AssessmentTemplate.is_active == True
    ).options(
        selectinload(AssessmentTemplate.dimensions),
        selectinload(AssessmentTemplate.questions).selectinload(AssessmentQuestion.dimension),
        selectinload(AssessmentTemplate.created_by)
    )

    result = await db.execute(query)
    template = result.scalar_one_or_none()

    if not template:
        return None

    return AssessmentTemplateDetailResponse.model_validate(template)


@router.get("/templates/{template_id}", response_model=AssessmentTemplateDetailResponse)
async def get_template(template_id: int, db: AsyncSession = Depends(get_db)):
    """Get a template with all its dimensions and questions."""
    query = select(AssessmentTemplate).where(
        AssessmentTemplate.id == template_id
    ).options(
        selectinload(AssessmentTemplate.dimensions),
        selectinload(AssessmentTemplate.questions).selectinload(AssessmentQuestion.dimension),
        selectinload(AssessmentTemplate.created_by)
    )

    result = await db.execute(query)
    template = result.scalar_one_or_none()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    return AssessmentTemplateDetailResponse.model_validate(template)


@router.post("/templates", response_model=AssessmentTemplateResponse, status_code=201)
async def create_template(
    template_in: AssessmentTemplateCreate,
    db: AsyncSession = Depends(get_db)
):
    """Create a new assessment template (JSON)."""
    # Create template
    template = AssessmentTemplate(
        name=template_in.name,
        version=template_in.version,
        description=template_in.description
    )
    db.add(template)
    await db.flush()

    # Create dimensions if provided
    dimension_map = {}  # To map dimension order to id for questions
    if template_in.dimensions:
        for i, dim_data in enumerate(template_in.dimensions):
            dimension = AssessmentDimension(
                template_id=template.id,
                name=dim_data.name,
                description=dim_data.description,
                display_order=dim_data.display_order or i,
                weight=dim_data.weight
            )
            db.add(dimension)
            await db.flush()
            dimension_map[dim_data.name] = dimension.id

    # Create questions if provided
    if template_in.questions:
        for i, q_data in enumerate(template_in.questions):
            question = AssessmentQuestion(
                template_id=template.id,
                dimension_id=q_data.dimension_id,
                question_text=q_data.question_text,
                question_number=q_data.question_number,
                min_score=q_data.min_score,
                max_score=q_data.max_score,
                score_labels=q_data.score_labels,
                display_order=q_data.display_order or i,
                is_required=q_data.is_required
            )
            db.add(question)

    await db.flush()

    return AssessmentTemplateResponse.model_validate(template)


@router.post("/templates/upload", response_model=ExcelUploadResult)
async def upload_template_excel(
    file: UploadFile = File(...),
    name: str = Query(..., description="Template name"),
    version: str = Query(..., description="Template version"),
    description: Optional[str] = Query(None, description="Template description"),
    db: AsyncSession = Depends(get_db)
):
    """Upload an Excel or CSV file to create a new assessment template."""
    errors = []
    contents = await file.read()
    filename = file.filename.lower() if file.filename else ""

    # Determine file type and parse accordingly
    if filename.endswith('.csv'):
        # Parse CSV file
        rows = parse_csv_file(contents)
    else:
        # Parse Excel file
        try:
            import openpyxl
            rows = parse_excel_file(contents)
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl library not installed. Please install it with: pip install openpyxl"
            )
        except Exception as e:
            return ExcelUploadResult(success=False, errors=[f"Failed to read Excel file: {str(e)}"])

    if not rows:
        return ExcelUploadResult(success=False, errors=["No data found in file"])

    # Check for SPM Maturity format (Domain, Lens, Question, Rating, Rating Label)
    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    is_spm_format = 'domain' in header and 'question' in header and 'rating' in header

    if is_spm_format:
        return await parse_spm_maturity_csv(rows, name, version, description, db)
    else:
        return await parse_tbm_format(rows, name, version, description, db)


@router.patch("/templates/{template_id}/update-ratings", response_model=ExcelUploadResult)
async def update_template_ratings(
    template_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Update rating descriptions and evidence for an existing template.

    Upload a CSV file with the same SPM format (Domain, Lens, Question, Rating, Rating Label,
    Rating Description, Evidence Required) to update the score_descriptions and score_evidence
    fields for matching questions. This preserves existing assessments while adding the missing data.
    """
    # Verify template exists
    template = await db.get(AssessmentTemplate, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # Parse the uploaded file
    contents = await file.read()
    filename = file.filename.lower() if file.filename else ""

    if filename.endswith('.csv'):
        rows = parse_csv_file(contents)
    else:
        try:
            rows = parse_excel_file(contents)
        except Exception as e:
            return ExcelUploadResult(success=False, errors=[f"Failed to read file: {str(e)}"])

    if not rows:
        return ExcelUploadResult(success=False, errors=["No data found in file"])

    # Find column indices from header
    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    try:
        domain_idx = header.index('domain')
        lens_idx = header.index('lens') if 'lens' in header else -1
        question_idx = header.index('question')
        rating_idx = header.index('rating')
    except ValueError as e:
        return ExcelUploadResult(success=False, errors=[f"Missing required column: {str(e)}"])

    # Check for rating description and evidence columns
    description_idx = header.index('rating description') if 'rating description' in header else -1
    evidence_idx = header.index('evidence required') if 'evidence required' in header else -1

    if description_idx == -1 and evidence_idx == -1:
        return ExcelUploadResult(success=False, errors=["No 'Rating Description' or 'Evidence Required' columns found in file"])

    # Build lookup of new data: {normalized_question_text: {rating: {description, evidence}}}
    update_data = {}
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or len(row) <= max(domain_idx, question_idx, rating_idx):
            continue

        domain = str(row[domain_idx]).strip() if row[domain_idx] else ""
        lens = str(row[lens_idx]).strip() if lens_idx >= 0 and len(row) > lens_idx and row[lens_idx] else ""
        question_text = str(row[question_idx]).strip() if row[question_idx] else ""
        rating = str(row[rating_idx]).strip() if row[rating_idx] else ""
        rating_desc = str(row[description_idx]).strip() if description_idx >= 0 and len(row) > description_idx and row[description_idx] else ""
        evidence = str(row[evidence_idx]).strip() if evidence_idx >= 0 and len(row) > evidence_idx and row[evidence_idx] else ""

        if not question_text or not rating:
            continue

        # Build the full question text as stored in the database
        full_question_text = f"[{lens}] {question_text}" if lens else question_text

        if full_question_text not in update_data:
            update_data[full_question_text] = {}

        update_data[full_question_text][rating] = {
            'description': rating_desc,
            'evidence': evidence
        }

    # Get all questions for this template
    result = await db.execute(
        select(AssessmentQuestion).where(AssessmentQuestion.template_id == template_id)
    )
    questions = result.scalars().all()

    questions_updated = 0
    errors = []

    # Build a normalized lookup for more flexible matching
    # Key: normalized question text (lowercase, stripped), Value: original key
    normalized_lookup = {}
    for key in update_data.keys():
        # Normalize: lowercase, strip whitespace
        normalized = key.lower().strip()
        normalized_lookup[normalized] = key
        # Also try without the lens prefix for matching
        if key.startswith('[') and '] ' in key:
            without_lens = key.split('] ', 1)[1].lower().strip()
            normalized_lookup[without_lens] = key

    for question in questions:
        rating_data = None

        # Try exact match first
        if question.question_text in update_data:
            rating_data = update_data[question.question_text]
        else:
            # Try normalized match
            normalized_q = question.question_text.lower().strip()
            if normalized_q in normalized_lookup:
                rating_data = update_data[normalized_lookup[normalized_q]]
            else:
                # Try matching without lens prefix
                if question.question_text.startswith('[') and '] ' in question.question_text:
                    without_lens = question.question_text.split('] ', 1)[1].lower().strip()
                    if without_lens in normalized_lookup:
                        rating_data = update_data[normalized_lookup[without_lens]]

        if rating_data:
            # Update score_descriptions
            new_descriptions = dict(question.score_descriptions) if question.score_descriptions else {}
            new_evidence = dict(question.score_evidence) if question.score_evidence else {}

            for rating, data in rating_data.items():
                if data['description']:
                    new_descriptions[rating] = data['description']
                if data['evidence']:
                    new_evidence[rating] = data['evidence']

            # Only update if there are changes
            if new_descriptions != question.score_descriptions or new_evidence != question.score_evidence:
                question.score_descriptions = new_descriptions
                question.score_evidence = new_evidence
                questions_updated += 1

    await db.commit()

    return ExcelUploadResult(
        success=True,
        template_id=template_id,
        questions_created=questions_updated,  # Reusing field to indicate questions updated
        errors=errors if errors else []
    )


async def parse_spm_maturity_csv(
    rows: List[List[str]],
    name: str,
    version: str,
    description: Optional[str],
    db: AsyncSession
) -> ExcelUploadResult:
    """Parse SPM Maturity Assessment CSV format (Domain, Lens, Question, Rating, Rating Label, Rating Description, Evidence Required)."""
    errors = []

    # Find column indices from header
    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    try:
        domain_idx = header.index('domain')
        lens_idx = header.index('lens')
        question_idx = header.index('question')
        rating_idx = header.index('rating')
        label_idx = header.index('rating label')
    except ValueError as e:
        return ExcelUploadResult(success=False, errors=[f"Missing required column: {str(e)}"])

    # Optional columns for rating description and evidence
    description_idx = header.index('rating description') if 'rating description' in header else -1
    evidence_idx = header.index('evidence required') if 'evidence required' in header else -1

    # Create template
    template = AssessmentTemplate(name=name, version=version, description=description)
    db.add(template)
    await db.flush()

    # Group questions by Domain + Lens + Question text (each question has 5 rows for ratings 1-5)
    # Structure: {domain: {(lens, question_text): {rating: {label, description, evidence}}}}
    question_data = {}

    for row_idx, row in enumerate(rows[1:], start=2):  # Skip header
        if not row or len(row) <= max(domain_idx, lens_idx, question_idx, rating_idx, label_idx):
            continue

        domain = str(row[domain_idx]).strip() if row[domain_idx] else ""
        lens = str(row[lens_idx]).strip() if row[lens_idx] else ""
        question_text = str(row[question_idx]).strip() if row[question_idx] else ""
        rating = str(row[rating_idx]).strip() if row[rating_idx] else ""
        label = str(row[label_idx]).strip() if row[label_idx] else ""
        rating_desc = str(row[description_idx]).strip() if description_idx >= 0 and len(row) > description_idx and row[description_idx] else ""
        evidence = str(row[evidence_idx]).strip() if evidence_idx >= 0 and len(row) > evidence_idx and row[evidence_idx] else ""

        if not domain or not question_text:
            continue

        if domain not in question_data:
            question_data[domain] = {}

        key = (lens, question_text)
        if key not in question_data[domain]:
            question_data[domain][key] = {}

        question_data[domain][key][rating] = {
            'label': label,
            'description': rating_desc,
            'evidence': evidence
        }

    # Create dimensions and questions
    dimension_map = {}
    dimensions_created = 0
    questions_created = 0

    for domain, questions in question_data.items():
        # Create dimension for this domain
        if domain.lower() not in dimension_map:
            dimension = AssessmentDimension(
                template_id=template.id,
                name=domain,
                description=None,
                display_order=dimensions_created
            )
            db.add(dimension)
            await db.flush()
            dimension_map[domain.lower()] = dimension.id
            dimensions_created += 1

        dim_id = dimension_map[domain.lower()]
        question_number = 1

        for (lens, question_text), rating_data in questions.items():
            # Build score labels, descriptions, and evidence dicts
            score_labels = {}
            score_descriptions = {}
            score_evidence = {}
            for rating, data in rating_data.items():
                if isinstance(data, dict):
                    score_labels[rating] = data.get('label', '')
                    if data.get('description'):
                        score_descriptions[rating] = data['description']
                    if data.get('evidence'):
                        score_evidence[rating] = data['evidence']
                else:
                    # Backwards compatibility if data is just the label string
                    score_labels[rating] = data

            # Determine min/max scores from the ratings we found
            ratings = [int(r) for r in score_labels.keys() if r.isdigit()]
            min_score = min(ratings) if ratings else 1
            max_score = max(ratings) if ratings else 5

            # Include lens in question text if present
            full_question_text = f"[{lens}] {question_text}" if lens else question_text

            question = AssessmentQuestion(
                template_id=template.id,
                dimension_id=dim_id,
                question_text=full_question_text,
                question_number=f"{dimensions_created}.{question_number}",
                min_score=min_score,
                max_score=max_score,
                score_labels=score_labels,
                score_descriptions=score_descriptions if score_descriptions else {},
                score_evidence=score_evidence if score_evidence else {},
                display_order=questions_created,
                is_required=True
            )
            db.add(question)
            questions_created += 1
            question_number += 1

    await db.flush()

    return ExcelUploadResult(
        success=True,
        template_id=template.id,
        dimensions_created=dimensions_created,
        questions_created=questions_created,
        errors=errors
    )


async def parse_tbm_format(
    rows: List[List[str]],
    name: str,
    version: str,
    description: Optional[str],
    db: AsyncSession
) -> ExcelUploadResult:
    """Parse TBM-style format (dimension headers with questions below)."""
    errors = []

    # Create template
    template = AssessmentTemplate(name=name, version=version, description=description)
    db.add(template)
    await db.flush()

    dimension_map = {}  # name -> dimension_id
    dimensions_created = 0
    questions_created = 0
    current_dimension = None
    current_dimension_desc = None
    question_number = 1

    for row_idx, row in enumerate(rows):
        if not row or not row[0]:
            continue

        first_cell = str(row[0]).strip()

        # Skip header row and empty rows
        if first_cell.lower().startswith('tbm practice') or first_cell.lower().startswith('maturity dimension'):
            continue

        # Check if this is a dimension header (single cell with no "0 - NA" in second column)
        second_cell = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        # If second cell is empty or doesn't look like a score option, this might be a dimension
        if not second_cell or (not second_cell.startswith('0 -') and not second_cell.startswith('Enter')):
            # Check if this is a short title (dimension name) - typically short and no punctuation at end
            if len(first_cell) < 50 and not first_cell.endswith('.') and not first_cell.endswith(','):
                # This is likely a dimension name
                current_dimension = first_cell
                current_dimension_desc = None
                question_number = 1
                continue
            else:
                # This might be a dimension description - store it
                if current_dimension and not current_dimension_desc:
                    current_dimension_desc = first_cell
                    continue

        # This is a question row
        if current_dimension and (second_cell.startswith('0 -') or 'NA' in second_cell.upper()):
            dim_key = current_dimension.lower()

            # Create dimension if not exists
            if dim_key not in dimension_map:
                dimension = AssessmentDimension(
                    template_id=template.id,
                    name=current_dimension,
                    description=current_dimension_desc,
                    display_order=dimensions_created
                )
                db.add(dimension)
                await db.flush()
                dimension_map[dim_key] = dimension.id
                dimensions_created += 1

            # Create question
            score_labels = {
                "0": "NA / Opt Out",
                "1": "Initial",
                "2": "Developing",
                "3": "Defined",
                "4": "Managed",
                "5": "Optimized"
            }

            question = AssessmentQuestion(
                template_id=template.id,
                dimension_id=dimension_map[dim_key],
                question_text=first_cell,
                question_number=f"{dimensions_created}.{question_number}",
                min_score=0,
                max_score=5,
                score_labels=score_labels,
                display_order=questions_created,
                is_required=True
            )
            db.add(question)
            questions_created += 1
            question_number += 1

    await db.flush()

    return ExcelUploadResult(
        success=True,
        template_id=template.id,
        dimensions_created=dimensions_created,
        questions_created=questions_created,
        errors=errors
    )


def parse_csv_file(contents: bytes) -> List[List[str]]:
    """Parse CSV file contents into rows."""
    try:
        # Try UTF-8 first, then fallback to latin-1
        try:
            text = contents.decode('utf-8')
        except UnicodeDecodeError:
            text = contents.decode('latin-1')

        reader = csv.reader(io.StringIO(text))
        return list(reader)
    except Exception as e:
        return []


def parse_excel_file(contents: bytes) -> List[List]:
    """Parse Excel file contents into rows."""
    import openpyxl
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active

    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


@router.patch("/templates/{template_id}", response_model=AssessmentTemplateResponse)
async def update_template(
    template_id: int,
    template_in: AssessmentTemplateUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update a template's metadata."""
    query = select(AssessmentTemplate).where(AssessmentTemplate.id == template_id)
    result = await db.execute(query)
    template = result.scalar_one_or_none()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    update_data = template_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(template, field, value)

    await db.flush()

    return AssessmentTemplateResponse.model_validate(template)


@router.post("/templates/{template_id}/activate", response_model=AssessmentTemplateResponse)
async def activate_template(template_id: int, db: AsyncSession = Depends(get_db)):
    """Set a template as the active version (deactivates others)."""
    # First deactivate all
    all_templates = await db.execute(select(AssessmentTemplate))
    for t in all_templates.scalars():
        t.is_active = False

    # Activate the specified one
    query = select(AssessmentTemplate).where(AssessmentTemplate.id == template_id)
    result = await db.execute(query)
    template = result.scalar_one_or_none()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    template.is_active = True
    await db.flush()
    await db.refresh(template)

    return AssessmentTemplateResponse.model_validate(template)


@router.delete("/templates/{template_id}", status_code=204)
async def delete_template(template_id: int, db: AsyncSession = Depends(get_db)):
    """Delete a template (only if no assessments use it)."""
    # Check if any assessments use this template
    count = await db.scalar(
        select(func.count()).select_from(CustomerAssessment).where(
            CustomerAssessment.template_id == template_id
        )
    )

    if count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete template: {count} assessments are using it"
        )

    query = select(AssessmentTemplate).where(AssessmentTemplate.id == template_id)
    result = await db.execute(query)
    template = result.scalar_one_or_none()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    await db.delete(template)


# ============================================================
# CUSTOMER ASSESSMENT ENDPOINTS
# ============================================================

@router.get("/customer/{customer_id}", response_model=CustomerAssessmentListResponse)
async def list_customer_assessments(
    customer_id: int,
    db: AsyncSession = Depends(get_db),
    status: Optional[AssessmentStatus] = None,
):
    """List all assessments for a customer."""
    query = select(CustomerAssessment).where(
        CustomerAssessment.customer_id == customer_id
    )

    if status:
        query = query.where(CustomerAssessment.status == status)

    query = query.order_by(CustomerAssessment.assessment_date.desc(), CustomerAssessment.id.desc())
    query = query.options(
        selectinload(CustomerAssessment.template),
        selectinload(CustomerAssessment.completed_by)
    )

    result = await db.execute(query)
    assessments = result.scalars().all()

    return CustomerAssessmentListResponse(
        items=[CustomerAssessmentResponse.model_validate(a) for a in assessments],
        total=len(assessments)
    )


@router.get("/customer/{customer_id}/history", response_model=AssessmentHistoryResponse)
async def get_customer_assessment_history(
    customer_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get assessment history with comparison to previous."""
    query = select(CustomerAssessment).where(
        and_(
            CustomerAssessment.customer_id == customer_id,
            CustomerAssessment.status == AssessmentStatus.COMPLETED
        )
    ).order_by(CustomerAssessment.assessment_date.desc(), CustomerAssessment.id.desc())
    query = query.options(selectinload(CustomerAssessment.template))

    result = await db.execute(query)
    assessments = list(result.scalars().all())

    comparison = None
    if len(assessments) >= 2:
        current = assessments[0]
        previous = assessments[1]

        # Calculate dimension changes
        dimension_changes = {}
        if current.dimension_scores and previous.dimension_scores:
            for dim in current.dimension_scores:
                if dim in previous.dimension_scores:
                    dimension_changes[dim] = round(
                        current.dimension_scores[dim] - previous.dimension_scores[dim], 2
                    )

        overall_change = None
        if current.overall_score is not None and previous.overall_score is not None:
            overall_change = round(current.overall_score - previous.overall_score, 2)

        comparison = AssessmentComparison(
            current=CustomerAssessmentResponse.model_validate(current),
            previous=CustomerAssessmentResponse.model_validate(previous),
            dimension_changes=dimension_changes,
            overall_change=overall_change
        )

    return AssessmentHistoryResponse(
        assessments=[CustomerAssessmentResponse.model_validate(a) for a in assessments],
        comparison=comparison
    )


@router.post("/customer/{customer_id}", response_model=CustomerAssessmentResponse, status_code=201)
async def create_customer_assessment(
    customer_id: int,
    assessment_in: CustomerAssessmentCreate,
    db: AsyncSession = Depends(get_db)
):
    """Start a new assessment for a customer."""
    # Verify template exists
    template = await db.get(AssessmentTemplate, assessment_in.template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    assessment = CustomerAssessment(
        customer_id=customer_id,
        template_id=assessment_in.template_id,
        assessment_date=assessment_in.assessment_date or date.today(),
        notes=assessment_in.notes,
        status=AssessmentStatus.DRAFT
    )
    db.add(assessment)
    await db.flush()

    return CustomerAssessmentResponse.model_validate(assessment)


@router.get("/{assessment_id}", response_model=CustomerAssessmentDetailResponse)
async def get_assessment(assessment_id: int, db: AsyncSession = Depends(get_db)):
    """Get an assessment with all responses."""
    query = select(CustomerAssessment).where(
        CustomerAssessment.id == assessment_id
    ).options(
        selectinload(CustomerAssessment.customer),
        selectinload(CustomerAssessment.template).selectinload(AssessmentTemplate.dimensions),
        selectinload(CustomerAssessment.template).selectinload(AssessmentTemplate.questions),
        selectinload(CustomerAssessment.completed_by),
        selectinload(CustomerAssessment.responses).selectinload(AssessmentResponse.question).selectinload(AssessmentQuestion.dimension)
    )

    result = await db.execute(query)
    assessment = result.scalar_one_or_none()

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    return CustomerAssessmentDetailResponse.model_validate(assessment)


@router.patch("/{assessment_id}", response_model=CustomerAssessmentResponse)
async def update_assessment(
    assessment_id: int,
    assessment_in: CustomerAssessmentUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update an assessment's status or notes."""
    query = select(CustomerAssessment).where(CustomerAssessment.id == assessment_id)
    result = await db.execute(query)
    assessment = result.scalar_one_or_none()

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    update_data = assessment_in.model_dump(exclude_unset=True)

    # If completing, set completed_at
    if update_data.get("status") == AssessmentStatus.COMPLETED:
        update_data["completed_at"] = datetime.utcnow()

    for field, value in update_data.items():
        setattr(assessment, field, value)

    await db.flush()

    return CustomerAssessmentResponse.model_validate(assessment)


@router.delete("/{assessment_id}", status_code=204)
async def delete_assessment(assessment_id: int, db: AsyncSession = Depends(get_db)):
    """Delete an assessment."""
    query = select(CustomerAssessment).where(CustomerAssessment.id == assessment_id)
    result = await db.execute(query)
    assessment = result.scalar_one_or_none()

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    await db.delete(assessment)


# ============================================================
# RESPONSE ENDPOINTS
# ============================================================

@router.post("/{assessment_id}/responses", response_model=CustomerAssessmentResponse)
async def save_responses(
    assessment_id: int,
    batch: BatchResponseSubmit,
    db: AsyncSession = Depends(get_db)
):
    """Save multiple responses at once."""
    query = select(CustomerAssessment).where(
        CustomerAssessment.id == assessment_id
    ).options(
        selectinload(CustomerAssessment.responses).selectinload(AssessmentResponse.question).selectinload(AssessmentQuestion.dimension)
    )
    result = await db.execute(query)
    assessment = result.scalar_one_or_none()

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    # Delete existing responses for questions being updated
    question_ids = {r.question_id for r in batch.responses}
    for existing in list(assessment.responses):
        if existing.question_id in question_ids:
            await db.delete(existing)

    # Add new responses
    for response_data in batch.responses:
        response = AssessmentResponse(
            customer_assessment_id=assessment_id,
            question_id=response_data.question_id,
            score=response_data.score,
            notes=response_data.notes
        )
        db.add(response)

    # Update status
    if batch.complete:
        assessment.status = AssessmentStatus.COMPLETED
        assessment.completed_at = datetime.utcnow()
        if batch.completed_by_id:
            assessment.completed_by_id = batch.completed_by_id
    elif assessment.status == AssessmentStatus.DRAFT:
        assessment.status = AssessmentStatus.IN_PROGRESS

    await db.flush()

    # Calculate scores using direct SQL query for reliability
    score_query = select(
        AssessmentDimension.name,
        func.avg(AssessmentResponse.score).label('avg_score')
    ).select_from(AssessmentResponse).join(
        AssessmentQuestion, AssessmentResponse.question_id == AssessmentQuestion.id
    ).join(
        AssessmentDimension, AssessmentQuestion.dimension_id == AssessmentDimension.id
    ).where(
        AssessmentResponse.customer_assessment_id == assessment_id
    ).group_by(AssessmentDimension.name)

    score_result = await db.execute(score_query)
    dimension_scores = {row.name: float(row.avg_score) for row in score_result}

    # Calculate overall score
    overall_score = sum(dimension_scores.values()) / len(dimension_scores) if dimension_scores else None

    # Update assessment with calculated scores
    assessment.dimension_scores = dimension_scores
    assessment.overall_score = overall_score

    await db.flush()
    await db.refresh(assessment)

    return CustomerAssessmentResponse.model_validate(assessment)


@router.post("/customer/{customer_id}/upload", response_model=ExcelResponseUploadResult)
async def upload_assessment_responses(
    customer_id: int,
    file: UploadFile = File(...),
    template_id: Optional[int] = Query(None, description="Template ID (uses active if not specified)"),
    assessment_date: Optional[date] = Query(None, description="Assessment date"),
    db: AsyncSession = Depends(get_db)
):
    """Upload an Excel file with completed assessment responses."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl library not installed"
        )

    errors = []

    # Get template
    if template_id:
        template = await db.get(AssessmentTemplate, template_id)
    else:
        query = select(AssessmentTemplate).where(AssessmentTemplate.is_active == True)
        result = await db.execute(query)
        template = result.scalar_one_or_none()

    if not template:
        return ExcelResponseUploadResult(
            success=False,
            errors=["No template specified and no active template found"]
        )

    # Load template questions
    query = select(AssessmentTemplate).where(
        AssessmentTemplate.id == template.id
    ).options(selectinload(AssessmentTemplate.questions))
    result = await db.execute(query)
    template = result.scalar_one()

    # Build question number -> question mapping
    question_map = {q.question_number: q for q in template.questions}

    # Read Excel file
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
    except Exception as e:
        return ExcelResponseUploadResult(success=False, errors=[f"Failed to read Excel file: {str(e)}"])

    # Create assessment
    assessment = CustomerAssessment(
        customer_id=customer_id,
        template_id=template.id,
        assessment_date=assessment_date or date.today(),
        status=AssessmentStatus.IN_PROGRESS
    )
    db.add(assessment)
    await db.flush()

    responses_saved = 0

    # Parse responses - expect columns like: Question Number, Score, Notes (optional)
    headers = [str(cell.value).lower() if cell.value else '' for cell in sheet[1]]
    num_col = next((i for i, h in enumerate(headers) if 'number' in h or 'num' in h or '#' in h or 'question' in h), 0)
    score_col = next((i for i, h in enumerate(headers) if 'score' in h or 'rating' in h or 'answer' in h), 1)
    notes_col = next((i for i, h in enumerate(headers) if 'note' in h or 'comment' in h), None)

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[num_col]:
            continue

        q_num = str(row[num_col]).strip()
        if q_num not in question_map:
            errors.append(f"Row {row_idx}: Unknown question number '{q_num}'")
            continue

        try:
            score = int(row[score_col])
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx}: Invalid score value")
            continue

        question = question_map[q_num]
        if score < question.min_score or score > question.max_score:
            errors.append(f"Row {row_idx}: Score {score} out of range ({question.min_score}-{question.max_score})")
            continue

        response = AssessmentResponse(
            customer_assessment_id=assessment.id,
            question_id=question.id,
            score=score,
            notes=str(row[notes_col]).strip() if notes_col and row[notes_col] else None
        )
        db.add(response)
        responses_saved += 1

    await db.flush()

    # Load responses and calculate scores
    query = select(CustomerAssessment).where(
        CustomerAssessment.id == assessment.id
    ).options(
        selectinload(CustomerAssessment.responses).selectinload(AssessmentResponse.question).selectinload(AssessmentQuestion.dimension)
    )
    result = await db.execute(query)
    assessment = result.scalar_one()

    # Calculate scores and mark complete
    assessment.calculate_scores()
    assessment.status = AssessmentStatus.COMPLETED
    assessment.completed_at = datetime.utcnow()
    await db.flush()

    return ExcelResponseUploadResult(
        success=True,
        assessment_id=assessment.id,
        responses_saved=responses_saved,
        errors=errors
    )


# ============================================================
# EXPORT ENDPOINTS
# ============================================================

@router.get("/{assessment_id}/export/excel")
async def export_assessment_excel(assessment_id: int, db: AsyncSession = Depends(get_db)):
    """Export assessment report to Excel."""
    # Get assessment with all related data
    query = select(CustomerAssessment).where(
        CustomerAssessment.id == assessment_id
    ).options(
        selectinload(CustomerAssessment.customer),
        selectinload(CustomerAssessment.template).selectinload(AssessmentTemplate.dimensions),
        selectinload(CustomerAssessment.completed_by),
        selectinload(CustomerAssessment.responses).selectinload(AssessmentResponse.question).selectinload(AssessmentQuestion.dimension),
        selectinload(CustomerAssessment.recommendations)
    )

    result = await db.execute(query)
    assessment = result.scalar_one_or_none()

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment Report"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    dimension_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header section
    ws['A1'] = "Assessment Report"
    ws['A1'].font = Font(bold=True, size=18)
    ws.merge_cells('A1:F1')

    # Customer and assessment info
    ws['A3'] = "Customer:"
    ws['B3'] = assessment.customer.name if assessment.customer else "N/A"
    ws['A3'].font = subheader_font

    ws['A4'] = "Template:"
    ws['B4'] = assessment.template.name if assessment.template else "N/A"
    ws['A4'].font = subheader_font

    ws['A5'] = "Assessment Date:"
    ws['B5'] = assessment.assessment_date.strftime("%Y-%m-%d") if assessment.assessment_date else "N/A"
    ws['A5'].font = subheader_font

    ws['A6'] = "Status:"
    ws['B6'] = assessment.status.value if assessment.status else "N/A"
    ws['A6'].font = subheader_font

    ws['A7'] = "Completed By:"
    if assessment.completed_by:
        ws['B7'] = f"{assessment.completed_by.first_name} {assessment.completed_by.last_name}"
    else:
        ws['B7'] = "N/A"
    ws['A7'].font = subheader_font

    ws['A8'] = "Completed At:"
    ws['B8'] = assessment.completed_at.strftime("%Y-%m-%d %H:%M") if assessment.completed_at else "N/A"
    ws['A8'].font = subheader_font

    ws['A9'] = "Overall Score:"
    ws['B9'] = f"{assessment.overall_score:.2f}" if assessment.overall_score else "N/A"
    ws['A9'].font = subheader_font
    ws['B9'].font = Font(bold=True, size=12)

    # Dimension scores summary
    ws['A11'] = "Dimension Scores"
    ws['A11'].font = header_font
    ws.merge_cells('A11:C11')

    row = 12
    if assessment.dimension_scores:
        for dim_name, score in assessment.dimension_scores.items():
            ws[f'A{row}'] = dim_name
            ws[f'B{row}'] = f"{score:.2f}"
            ws[f'A{row}'].font = subheader_font
            row += 1

    # Questions and responses header
    row += 2
    header_row = row
    headers = ["#", "Dimension", "Question", "Score", "Label", "Rating Description", "Evidence Required", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40

    # Build response lookup
    response_lookup = {r.question_id: r for r in assessment.responses}

    # Get all questions from template, sorted by dimension and display order
    questions = sorted(
        assessment.template.questions if assessment.template else [],
        key=lambda q: (q.dimension.display_order if q.dimension else 0, q.display_order)
    )

    # Add question rows
    row += 1
    current_dimension = None
    for question in questions:
        response = response_lookup.get(question.id)

        # Add dimension separator
        dim_name = question.dimension.name if question.dimension else "General"
        if dim_name != current_dimension:
            current_dimension = dim_name

        # Question number
        ws.cell(row=row, column=1, value=question.question_number).border = thin_border

        # Dimension
        ws.cell(row=row, column=2, value=dim_name).border = thin_border

        # Question text
        q_cell = ws.cell(row=row, column=3, value=question.question_text)
        q_cell.border = thin_border
        q_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Score
        score_cell = ws.cell(row=row, column=4, value=response.score if response else "")
        score_cell.border = thin_border
        score_cell.alignment = Alignment(horizontal='center')

        # Score label
        score_label = ""
        score_description = ""
        score_evidence = ""
        if response:
            score_key = str(response.score)
            if question.score_labels:
                score_label = question.score_labels.get(score_key, "")
            if question.score_descriptions:
                score_description = question.score_descriptions.get(score_key, "")
            if question.score_evidence:
                score_evidence = question.score_evidence.get(score_key, "")

        ws.cell(row=row, column=5, value=score_label).border = thin_border

        # Rating Description
        desc_cell = ws.cell(row=row, column=6, value=score_description)
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Evidence Required
        evidence_cell = ws.cell(row=row, column=7, value=score_evidence)
        evidence_cell.border = thin_border
        evidence_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Notes
        notes_cell = ws.cell(row=row, column=8, value=response.notes if response else "")
        notes_cell.border = thin_border
        notes_cell.alignment = Alignment(wrap_text=True, vertical='top')

        row += 1

    # Add Recommendations section if there are any
    if assessment.recommendations:
        row += 2  # Add spacing

        # Recommendations header
        ws[f'A{row}'] = "Recommendations"
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        # Recommendations table header
        rec_headers = ["#", "Priority", "Category", "Title", "Description", "", "", "Created By"]
        for col, header in enumerate(rec_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        # Add recommendation rows
        for idx, rec in enumerate(sorted(assessment.recommendations, key=lambda r: r.display_order), 1):
            # Number
            ws.cell(row=row, column=1, value=idx).border = thin_border

            # Priority
            priority_cell = ws.cell(row=row, column=2, value=rec.priority.value.upper() if rec.priority else "MEDIUM")
            priority_cell.border = thin_border
            priority_cell.alignment = Alignment(horizontal='center')
            # Color code priority
            if rec.priority:
                if rec.priority.value == "high":
                    priority_cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                elif rec.priority.value == "medium":
                    priority_cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                elif rec.priority.value == "low":
                    priority_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

            # Category
            ws.cell(row=row, column=3, value=rec.category or "").border = thin_border

            # Title
            title_cell = ws.cell(row=row, column=4, value=rec.title)
            title_cell.border = thin_border
            title_cell.font = Font(bold=True)

            # Description (merge columns 5-7)
            desc_cell = ws.cell(row=row, column=5, value=rec.description or "")
            desc_cell.border = thin_border
            desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)

            # Created By
            ws.cell(row=row, column=8, value=rec.created_by or "").border = thin_border

            row += 1

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    customer_name = assessment.customer.name.replace(" ", "_") if assessment.customer else "assessment"
    filename = f"{customer_name}_assessment_{assessment.assessment_date or 'draft'}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{assessment_id}/report")
async def get_assessment_report(assessment_id: int, db: AsyncSession = Depends(get_db)):
    """Get assessment report data for display."""
    query = select(CustomerAssessment).where(
        CustomerAssessment.id == assessment_id
    ).options(
        selectinload(CustomerAssessment.customer),
        selectinload(CustomerAssessment.template).selectinload(AssessmentTemplate.dimensions),
        selectinload(CustomerAssessment.template).selectinload(AssessmentTemplate.questions).selectinload(AssessmentQuestion.dimension),
        selectinload(CustomerAssessment.completed_by),
        selectinload(CustomerAssessment.responses).selectinload(AssessmentResponse.question).selectinload(AssessmentQuestion.dimension)
    )

    result = await db.execute(query)
    assessment = result.scalar_one_or_none()

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    # Build response lookup
    response_lookup = {r.question_id: r for r in assessment.responses}

    # Build report data grouped by dimension
    dimensions_data = []
    questions_by_dimension = {}

    for question in assessment.template.questions if assessment.template else []:
        dim_name = question.dimension.name if question.dimension else "General"
        if dim_name not in questions_by_dimension:
            questions_by_dimension[dim_name] = {
                "dimension_id": question.dimension.id if question.dimension else None,
                "dimension_name": dim_name,
                "display_order": question.dimension.display_order if question.dimension else 999,
                "questions": []
            }

        response = response_lookup.get(question.id)
        score_label = ""
        score_description = ""
        score_evidence = ""
        if response:
            score_key = str(response.score)
            if question.score_labels:
                score_label = question.score_labels.get(score_key, "")
            if question.score_descriptions:
                score_description = question.score_descriptions.get(score_key, "")
            if question.score_evidence:
                score_evidence = question.score_evidence.get(score_key, "")

        questions_by_dimension[dim_name]["questions"].append({
            "question_id": question.id,
            "question_number": question.question_number,
            "question_text": question.question_text,
            "score": response.score if response else None,
            "score_label": score_label,
            "score_description": score_description,
            "score_evidence": score_evidence,
            "notes": response.notes if response else None,
            "min_score": question.min_score,
            "max_score": question.max_score
        })

    # Sort dimensions by display order
    dimensions_data = sorted(questions_by_dimension.values(), key=lambda d: d["display_order"])

    # Sort questions within each dimension
    for dim in dimensions_data:
        dim["questions"].sort(key=lambda q: q["question_number"])

    # Fetch roadmap recommendations for this assessment (not dismissed)
    rec_query = select(RoadmapRecommendation).where(
        RoadmapRecommendation.customer_assessment_id == assessment_id,
        RoadmapRecommendation.is_dismissed == False
    ).options(
        selectinload(RoadmapRecommendation.use_case),
        selectinload(RoadmapRecommendation.tp_feature_mapping)
    ).order_by(RoadmapRecommendation.priority_score.desc())

    rec_result = await db.execute(rec_query)
    roadmap_recommendations = rec_result.scalars().all()

    # Build recommendations data from roadmap recommendations
    recommendations_data = []
    for rec in roadmap_recommendations:
        recommendations_data.append({
            "id": rec.id,
            "title": rec.title,
            "description": rec.description,
            "dimension_name": rec.dimension_name,
            "dimension_score": rec.dimension_score,
            "priority_score": rec.priority_score,
            "improvement_potential": rec.improvement_potential,
            "is_accepted": rec.is_accepted,
            "use_case_name": rec.use_case.name if rec.use_case else None,
            "solution_area": rec.use_case.solution_area if rec.use_case else None,
            "tp_feature_name": rec.tp_feature_mapping.tp_feature_name if rec.tp_feature_mapping else None,
            "tp_feature_id": rec.tp_feature_mapping.tp_feature_id if rec.tp_feature_mapping else None,
            "created_at": rec.created_at.isoformat() if rec.created_at else None
        })

    return {
        "assessment_id": assessment.id,
        "customer": {
            "id": assessment.customer.id if assessment.customer else None,
            "name": assessment.customer.name if assessment.customer else "N/A"
        },
        "template": {
            "id": assessment.template.id if assessment.template else None,
            "name": assessment.template.name if assessment.template else "N/A",
            "version": assessment.template.version if assessment.template else "N/A"
        },
        "assessment_date": assessment.assessment_date.isoformat() if assessment.assessment_date else None,
        "status": assessment.status.value if assessment.status else None,
        "overall_score": assessment.overall_score,
        "dimension_scores": assessment.dimension_scores or {},
        "completed_by": {
            "id": assessment.completed_by.id if assessment.completed_by else None,
            "name": f"{assessment.completed_by.first_name} {assessment.completed_by.last_name}" if assessment.completed_by else None
        },
        "completed_at": assessment.completed_at.isoformat() if assessment.completed_at else None,
        "notes": assessment.notes,
        "dimensions": dimensions_data,
        "recommendations": recommendations_data,
        "total_questions": len(assessment.template.questions) if assessment.template else 0,
        "answered_questions": len(assessment.responses)
    }


# ============================================================
# RESPONSE EDITING ENDPOINTS
# ============================================================

@router.patch("/{assessment_id}/responses/{response_id}", response_model=AssessmentAnswerResponse)
async def update_response(
    assessment_id: int,
    response_id: int,
    update_data: AssessmentAnswerUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Edit a response score or notes with audit trail."""
    # Get the response
    query = select(AssessmentResponse).where(
        and_(
            AssessmentResponse.id == response_id,
            AssessmentResponse.customer_assessment_id == assessment_id
        )
    ).options(selectinload(AssessmentResponse.question))
    result = await db.execute(query)
    response = result.scalar_one_or_none()

    if not response:
        raise HTTPException(status_code=404, detail="Response not found")

    # Track changes for audit
    audit_entries = []

    if update_data.score is not None and update_data.score != response.score:
        audit_entries.append(AssessmentResponseAudit(
            response_id=response.id,
            customer_assessment_id=assessment_id,
            question_id=response.question_id,
            field_changed="score",
            old_value=str(response.score),
            new_value=str(update_data.score),
            change_reason=update_data.change_reason,
            changed_by_id=update_data.edited_by_id
        ))
        response.score = update_data.score

    if update_data.notes is not None and update_data.notes != response.notes:
        audit_entries.append(AssessmentResponseAudit(
            response_id=response.id,
            customer_assessment_id=assessment_id,
            question_id=response.question_id,
            field_changed="notes",
            old_value=response.notes,
            new_value=update_data.notes,
            change_reason=update_data.change_reason,
            changed_by_id=update_data.edited_by_id
        ))
        response.notes = update_data.notes

    # Update edit tracking
    if audit_entries:
        response.last_edited_at = datetime.utcnow()
        response.last_edited_by_id = update_data.edited_by_id
        for entry in audit_entries:
            db.add(entry)

    await db.flush()

    # Recalculate assessment scores
    await recalculate_assessment_scores(assessment_id, db)

    return AssessmentAnswerResponse.model_validate(response)


async def recalculate_assessment_scores(assessment_id: int, db: AsyncSession):
    """Recalculate dimension and overall scores for an assessment."""
    score_query = select(
        AssessmentDimension.name,
        func.avg(AssessmentResponse.score).label('avg_score')
    ).select_from(AssessmentResponse).join(
        AssessmentQuestion, AssessmentResponse.question_id == AssessmentQuestion.id
    ).join(
        AssessmentDimension, AssessmentQuestion.dimension_id == AssessmentDimension.id
    ).where(
        AssessmentResponse.customer_assessment_id == assessment_id
    ).group_by(AssessmentDimension.name)

    score_result = await db.execute(score_query)
    dimension_scores = {row.name: float(row.avg_score) for row in score_result}

    # Calculate overall score
    overall_score = sum(dimension_scores.values()) / len(dimension_scores) if dimension_scores else None

    # Update assessment
    assessment = await db.get(CustomerAssessment, assessment_id)
    if assessment:
        assessment.dimension_scores = dimension_scores
        assessment.overall_score = overall_score
        await db.flush()


# ============================================================
# AUDIT TRAIL ENDPOINTS
# ============================================================

@router.get("/{assessment_id}/audit", response_model=AssessmentAuditListResponse)
async def get_assessment_audit(
    assessment_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get audit trail for all changes made to an assessment's responses."""
    query = select(AssessmentResponseAudit).where(
        AssessmentResponseAudit.customer_assessment_id == assessment_id
    ).options(
        selectinload(AssessmentResponseAudit.changed_by),
        selectinload(AssessmentResponseAudit.question)
    ).order_by(AssessmentResponseAudit.changed_at.desc())

    result = await db.execute(query)
    audit_entries = result.scalars().all()

    return AssessmentAuditListResponse(
        items=[AssessmentAuditEntry.model_validate(entry) for entry in audit_entries],
        total=len(audit_entries)
    )


# ============================================================
# TARGET ENDPOINTS
# ============================================================

@router.get("/customer/{customer_id}/targets", response_model=TargetListResponse)
async def list_customer_targets(
    customer_id: int,
    active_only: bool = Query(True, description="Only return active targets"),
    db: AsyncSession = Depends(get_db)
):
    """List all assessment targets for a customer."""
    query = select(CustomerAssessmentTarget).where(
        CustomerAssessmentTarget.customer_id == customer_id
    )

    if active_only:
        query = query.where(CustomerAssessmentTarget.is_active == True)

    query = query.order_by(CustomerAssessmentTarget.target_date.desc().nullslast())
    query = query.options(selectinload(CustomerAssessmentTarget.created_by))

    result = await db.execute(query)
    targets = result.scalars().all()

    return TargetListResponse(
        items=[TargetResponse.model_validate(t) for t in targets],
        total=len(targets)
    )


@router.post("/customer/{customer_id}/targets", response_model=TargetResponse, status_code=201)
async def create_customer_target(
    customer_id: int,
    target_in: TargetCreate,
    db: AsyncSession = Depends(get_db)
):
    """Create a new assessment target for a customer."""
    # Calculate overall target if not provided
    overall_target = target_in.overall_target
    if overall_target is None and target_in.target_scores:
        overall_target = sum(target_in.target_scores.values()) / len(target_in.target_scores)

    target = CustomerAssessmentTarget(
        customer_id=customer_id,
        name=target_in.name,
        description=target_in.description,
        target_date=target_in.target_date,
        target_scores=target_in.target_scores,
        overall_target=overall_target,
        is_active=target_in.is_active,
        created_by_id=target_in.created_by_id
    )
    db.add(target)
    await db.flush()
    await db.refresh(target)

    return TargetResponse.model_validate(target)


@router.get("/targets/{target_id}", response_model=TargetResponse)
async def get_target(
    target_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get a specific target."""
    query = select(CustomerAssessmentTarget).where(
        CustomerAssessmentTarget.id == target_id
    ).options(selectinload(CustomerAssessmentTarget.created_by))

    result = await db.execute(query)
    target = result.scalar_one_or_none()

    if not target:
        raise HTTPException(status_code=404, detail="Target not found")

    return TargetResponse.model_validate(target)


@router.patch("/targets/{target_id}", response_model=TargetResponse)
async def update_target(
    target_id: int,
    target_in: TargetUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update an existing target."""
    query = select(CustomerAssessmentTarget).where(
        CustomerAssessmentTarget.id == target_id
    )
    result = await db.execute(query)
    target = result.scalar_one_or_none()

    if not target:
        raise HTTPException(status_code=404, detail="Target not found")

    update_data = target_in.model_dump(exclude_unset=True)

    # Recalculate overall target if target_scores changed
    if 'target_scores' in update_data and update_data['target_scores']:
        if 'overall_target' not in update_data:
            update_data['overall_target'] = sum(update_data['target_scores'].values()) / len(update_data['target_scores'])

    for field, value in update_data.items():
        setattr(target, field, value)

    await db.flush()
    await db.refresh(target)

    return TargetResponse.model_validate(target)


@router.delete("/targets/{target_id}", status_code=204)
async def delete_target(
    target_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Delete a target."""
    query = select(CustomerAssessmentTarget).where(
        CustomerAssessmentTarget.id == target_id
    )
    result = await db.execute(query)
    target = result.scalar_one_or_none()

    if not target:
        raise HTTPException(status_code=404, detail="Target not found")

    await db.delete(target)


@router.get("/customer/{customer_id}/targets/{target_id}/gap-analysis", response_model=GapAnalysisResponse)
async def get_gap_analysis(
    customer_id: int,
    target_id: int,
    assessment_id: Optional[int] = Query(None, description="Specific assessment to compare against"),
    db: AsyncSession = Depends(get_db)
):
    """Get gap analysis between a target and current/specified assessment."""
    # Get target
    target_query = select(CustomerAssessmentTarget).where(
        and_(
            CustomerAssessmentTarget.id == target_id,
            CustomerAssessmentTarget.customer_id == customer_id
        )
    ).options(selectinload(CustomerAssessmentTarget.created_by))

    target_result = await db.execute(target_query)
    target = target_result.scalar_one_or_none()

    if not target:
        raise HTTPException(status_code=404, detail="Target not found")

    # Get assessment (most recent completed if not specified)
    if assessment_id:
        assessment_query = select(CustomerAssessment).where(
            CustomerAssessment.id == assessment_id
        )
    else:
        assessment_query = select(CustomerAssessment).where(
            and_(
                CustomerAssessment.customer_id == customer_id,
                CustomerAssessment.status == AssessmentStatus.COMPLETED
            )
        ).order_by(CustomerAssessment.assessment_date.desc(), CustomerAssessment.id.desc())

    assessment_result = await db.execute(assessment_query)
    assessment = assessment_result.scalar_one_or_none()

    # Calculate dimension gaps
    dimension_gaps = []
    current_scores = assessment.dimension_scores if assessment else {}

    for dim_name, target_score in (target.target_scores or {}).items():
        current_score = current_scores.get(dim_name)
        gap = None
        status = "at_risk"

        if current_score is not None:
            gap = target_score - current_score
            if gap <= 0:
                status = "achieved"
            elif gap <= 0.5:
                status = "on_track"
            elif gap <= 1.0:
                status = "needs_attention"
            else:
                status = "at_risk"

        dimension_gaps.append(DimensionGap(
            dimension_name=dim_name,
            current_score=current_score,
            target_score=target_score,
            gap=gap,
            status=status
        ))

    # Calculate overall gap
    current_overall = assessment.overall_score if assessment else None
    target_overall = target.overall_target
    overall_gap = None
    overall_status = "at_risk"

    if current_overall is not None and target_overall is not None:
        overall_gap = target_overall - current_overall
        if overall_gap <= 0:
            overall_status = "achieved"
        elif overall_gap <= 0.5:
            overall_status = "on_track"
        elif overall_gap <= 1.0:
            overall_status = "needs_attention"
        else:
            overall_status = "at_risk"

    # Calculate days to target
    days_to_target = None
    if target.target_date:
        days_to_target = (target.target_date - date.today()).days

    return GapAnalysisResponse(
        target=TargetResponse.model_validate(target),
        current_overall=current_overall,
        target_overall=target_overall,
        overall_gap=overall_gap,
        overall_status=overall_status,
        dimension_gaps=dimension_gaps,
        days_to_target=days_to_target
    )


# ============================================================
# FLOW VISUALIZATION ENDPOINT
# ============================================================

@router.get("/customer/{customer_id}/flow-visualization", response_model=FlowVisualizationResponse)
async def get_flow_visualization(
    customer_id: int,
    threshold: float = Query(3.5, ge=1.0, le=5.0, description="Score threshold for weak dimensions"),
    db: AsyncSession = Depends(get_db)
):
    """
    Get flow visualization data showing:
    Assessment dimensions (with gaps) -> Recommended use cases -> TP solutions to implement.

    This provides data for a Sankey diagram visualization.
    """
    from app.models.mapping import DimensionUseCaseMapping
    from app.models.use_case import CustomerUseCase, UseCaseStatus
    from app.models.use_case_solution_mapping import UseCaseTPSolutionMapping
    from app.models.tp_solution import TPSolution

    # 1. Get customer's latest completed assessment
    assessment_query = select(CustomerAssessment).where(
        and_(
            CustomerAssessment.customer_id == customer_id,
            CustomerAssessment.status == AssessmentStatus.COMPLETED
        )
    ).order_by(CustomerAssessment.completed_at.desc()).limit(1)

    result = await db.execute(assessment_query)
    assessment = result.scalar_one_or_none()

    if not assessment or not assessment.dimension_scores:
        return FlowVisualizationResponse(
            customer_id=customer_id,
            assessment_id=None,
            nodes=[],
            links=[],
            weak_dimensions_count=0,
            recommended_use_cases_count=0,
            tp_solutions_count=0
        )

    # 2. Identify weak dimensions (below threshold)
    weak_dims = []
    for dim_name, score in assessment.dimension_scores.items():
        if score < threshold:
            gap = threshold - score
            weak_dims.append((dim_name, score, gap))

    # Sort by gap (worst first)
    weak_dims.sort(key=lambda x: x[2], reverse=True)

    if not weak_dims:
        return FlowVisualizationResponse(
            customer_id=customer_id,
            assessment_id=assessment.id,
            nodes=[],
            links=[],
            weak_dimensions_count=0,
            recommended_use_cases_count=0,
            tp_solutions_count=0
        )

    weak_dim_names = [wd[0] for wd in weak_dims]
    weak_dim_lookup = {wd[0]: (wd[1], wd[2]) for wd in weak_dims}

    # 3. Get dimension IDs from names
    dim_query = select(AssessmentDimension).where(
        AssessmentDimension.name.in_(weak_dim_names)
    )
    result = await db.execute(dim_query)
    dimensions = result.scalars().all()
    dim_id_to_name = {d.id: d.name for d in dimensions}
    dim_ids = list(dim_id_to_name.keys())

    if not dim_ids:
        return FlowVisualizationResponse(
            customer_id=customer_id,
            assessment_id=assessment.id,
            nodes=[],
            links=[],
            weak_dimensions_count=len(weak_dims),
            recommended_use_cases_count=0,
            tp_solutions_count=0
        )

    # 4. Get use cases mapped to weak dimensions
    mapping_query = select(DimensionUseCaseMapping).where(
        DimensionUseCaseMapping.dimension_id.in_(dim_ids)
    ).options(
        selectinload(DimensionUseCaseMapping.dimension),
        selectinload(DimensionUseCaseMapping.use_case)
    )

    result = await db.execute(mapping_query)
    dim_use_case_mappings = result.scalars().all()

    # 5. Filter out already-implemented use cases
    implemented_query = select(CustomerUseCase.use_case_id).where(
        CustomerUseCase.customer_id == customer_id,
        CustomerUseCase.status.in_([
            UseCaseStatus.IMPLEMENTED,
            UseCaseStatus.OPTIMIZED,
            UseCaseStatus.IN_PROGRESS
        ])
    )
    result = await db.execute(implemented_query)
    implemented_ids = set(result.scalars().all())

    candidate_mappings = [
        m for m in dim_use_case_mappings
        if m.use_case_id not in implemented_ids
    ]

    if not candidate_mappings:
        # Return just the dimension nodes
        nodes = []
        for dim_name, score, gap in weak_dims:
            nodes.append(FlowNode(
                id=f"dim_{dim_name}",
                name=dim_name,
                type="dimension",
                score=score,
                gap=gap
            ))

        return FlowVisualizationResponse(
            customer_id=customer_id,
            assessment_id=assessment.id,
            nodes=nodes,
            links=[],
            weak_dimensions_count=len(weak_dims),
            recommended_use_cases_count=0,
            tp_solutions_count=0
        )

    # 6. Get TP solutions for candidate use cases
    use_case_ids = list(set(m.use_case_id for m in candidate_mappings))

    tp_solution_query = select(UseCaseTPSolutionMapping).where(
        UseCaseTPSolutionMapping.use_case_id.in_(use_case_ids)
    ).options(
        selectinload(UseCaseTPSolutionMapping.use_case),
        selectinload(UseCaseTPSolutionMapping.tp_solution)
    )

    result = await db.execute(tp_solution_query)
    tp_solution_mappings = result.scalars().all()

    # Group TP solutions by use case
    tp_by_use_case = {}
    for mapping in tp_solution_mappings:
        if mapping.use_case_id not in tp_by_use_case:
            tp_by_use_case[mapping.use_case_id] = []
        tp_by_use_case[mapping.use_case_id].append(mapping)

    # 7. Build nodes and links
    nodes = []
    links = []
    seen_dims = set()
    seen_use_cases = set()
    seen_tp_solutions = set()

    # Add dimension nodes
    for dim_name, score, gap in weak_dims:
        dim_id = f"dim_{dim_name}"
        if dim_id not in seen_dims:
            nodes.append(FlowNode(
                id=dim_id,
                name=dim_name,
                type="dimension",
                score=score,
                gap=gap
            ))
            seen_dims.add(dim_id)

    # Add use case nodes and dimension->use case links
    for mapping in candidate_mappings:
        if not mapping.use_case:
            continue

        uc_id = f"uc_{mapping.use_case_id}"
        dim_name = mapping.dimension.name if mapping.dimension else "Unknown"
        dim_id = f"dim_{dim_name}"

        # Add use case node if not already added
        if uc_id not in seen_use_cases:
            nodes.append(FlowNode(
                id=uc_id,
                name=mapping.use_case.name,
                type="use_case",
                solution_area=mapping.use_case.solution_area
            ))
            seen_use_cases.add(uc_id)

        # Add dimension -> use case link
        dim_score, dim_gap = weak_dim_lookup.get(dim_name, (0, 0))
        links.append(FlowLink(
            source=dim_id,
            target=uc_id,
            value=mapping.impact_weight,
            impact_weight=mapping.impact_weight
        ))

    # Add TP solution nodes and use case->TP solution links
    for uc_id_num, solution_list in tp_by_use_case.items():
        uc_id = f"uc_{uc_id_num}"

        for mapping in solution_list:
            if not mapping.tp_solution:
                continue

            tp_node_id = f"tp_{mapping.tp_solution_id}"

            # Add TP solution node if not already added
            if tp_node_id not in seen_tp_solutions:
                nodes.append(FlowNode(
                    id=tp_node_id,
                    name=mapping.tp_solution.name,
                    type="tp_solution",
                    tp_id=mapping.tp_solution_id,
                    is_required=mapping.is_required,
                    category=mapping.tp_solution.category.value if mapping.tp_solution.category else None,
                    version=mapping.tp_solution.version
                ))
                seen_tp_solutions.add(tp_node_id)

            # Add use case -> TP solution link
            links.append(FlowLink(
                source=uc_id,
                target=tp_node_id,
                value=1.0 if mapping.is_required else 0.5,
                is_required=mapping.is_required
            ))

    return FlowVisualizationResponse(
        customer_id=customer_id,
        assessment_id=assessment.id,
        nodes=nodes,
        links=links,
        weak_dimensions_count=len(weak_dims),
        recommended_use_cases_count=len(seen_use_cases),
        tp_solutions_count=len(seen_tp_solutions)
    )


# ============================================================
# RECOMMENDATION ENDPOINTS
# ============================================================

@router.get("/{assessment_id}/recommendations", response_model=AssessmentRecommendationListResponse)
async def list_assessment_recommendations(
    assessment_id: int,
    db: AsyncSession = Depends(get_db)
):
    """List all recommendations for an assessment."""
    # Verify assessment exists
    assessment = await db.get(CustomerAssessment, assessment_id)
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    query = select(AssessmentRecommendation).where(
        AssessmentRecommendation.assessment_id == assessment_id
    ).order_by(AssessmentRecommendation.display_order, AssessmentRecommendation.created_at)

    result = await db.execute(query)
    recommendations = result.scalars().all()

    return AssessmentRecommendationListResponse(
        items=[AssessmentRecommendationResponse.model_validate(r) for r in recommendations],
        total=len(recommendations)
    )


@router.post("/{assessment_id}/recommendations", response_model=AssessmentRecommendationResponse, status_code=201)
async def create_assessment_recommendation(
    assessment_id: int,
    recommendation_in: AssessmentRecommendationCreate,
    db: AsyncSession = Depends(get_db)
):
    """Create a new recommendation for an assessment."""
    # Verify assessment exists
    assessment = await db.get(CustomerAssessment, assessment_id)
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    # Get max display_order for this assessment
    max_order_query = select(func.max(AssessmentRecommendation.display_order)).where(
        AssessmentRecommendation.assessment_id == assessment_id
    )
    result = await db.execute(max_order_query)
    max_order = result.scalar() or 0

    recommendation = AssessmentRecommendation(
        assessment_id=assessment_id,
        title=recommendation_in.title,
        description=recommendation_in.description,
        priority=recommendation_in.priority,
        category=recommendation_in.category,
        display_order=recommendation_in.display_order if recommendation_in.display_order > 0 else max_order + 1,
        created_by=recommendation_in.created_by
    )
    db.add(recommendation)
    await db.flush()
    await db.refresh(recommendation)

    return AssessmentRecommendationResponse.model_validate(recommendation)


@router.get("/{assessment_id}/recommendations/{recommendation_id}", response_model=AssessmentRecommendationResponse)
async def get_assessment_recommendation(
    assessment_id: int,
    recommendation_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get a specific recommendation."""
    query = select(AssessmentRecommendation).where(
        and_(
            AssessmentRecommendation.id == recommendation_id,
            AssessmentRecommendation.assessment_id == assessment_id
        )
    )
    result = await db.execute(query)
    recommendation = result.scalar_one_or_none()

    if not recommendation:
        raise HTTPException(status_code=404, detail="Recommendation not found")

    return AssessmentRecommendationResponse.model_validate(recommendation)


@router.patch("/{assessment_id}/recommendations/{recommendation_id}", response_model=AssessmentRecommendationResponse)
async def update_assessment_recommendation(
    assessment_id: int,
    recommendation_id: int,
    recommendation_in: AssessmentRecommendationUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update an existing recommendation."""
    query = select(AssessmentRecommendation).where(
        and_(
            AssessmentRecommendation.id == recommendation_id,
            AssessmentRecommendation.assessment_id == assessment_id
        )
    )
    result = await db.execute(query)
    recommendation = result.scalar_one_or_none()

    if not recommendation:
        raise HTTPException(status_code=404, detail="Recommendation not found")

    update_data = recommendation_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(recommendation, field, value)

    await db.flush()
    await db.refresh(recommendation)

    return AssessmentRecommendationResponse.model_validate(recommendation)


@router.delete("/{assessment_id}/recommendations/{recommendation_id}", status_code=204)
async def delete_assessment_recommendation(
    assessment_id: int,
    recommendation_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Delete a recommendation."""
    query = select(AssessmentRecommendation).where(
        and_(
            AssessmentRecommendation.id == recommendation_id,
            AssessmentRecommendation.assessment_id == assessment_id
        )
    )
    result = await db.execute(query)
    recommendation = result.scalar_one_or_none()

    if not recommendation:
        raise HTTPException(status_code=404, detail="Recommendation not found")

    await db.delete(recommendation)
